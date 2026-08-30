#!/usr/bin/env python3
"""Build a tracking-plan workbook from a single JSON file.

    python3 build_workbook.py plan.json [-o output.xlsx]

Every tab is generated from the JSON, and the tables that could drift (registered
dimensions, the user-property list) are derived from the source rather than kept by
hand, because a hand-maintained copy silently falls out of step and nobody notices
until an analytics tool has been configured from the stale version.

Screenshots named SCR-nn_*.png in a `tracking_plan_screens/` folder beside the plan
are embedded as thumbnails on a gallery tab, and the per-event screen references
become internal links to that tab. Internal links are used deliberately: local file
links fail in most spreadsheet apps and always in browser-based ones.

Only `meta` and `events` are required. Every other section is optional and its tab is
skipped when absent, so the same script serves a five-event first pass and a
two-hundred-event mature plan.
"""

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

SCREEN_DIR = "tracking_plan_screens"
GALLERY_TAB = "Screen gallery"

# Palette. Override any of these under meta.theme in the plan JSON.
THEME = {
    "ink": "2B1F17",      # header fills and body text
    "accent": "C2410C",   # event names, links, section headings
    "paper": "FAF7F2",    # row banding
    "grey": "6B6259",     # secondary text
    "line": "D8D2C9",     # cell borders
    "font": "Arial",
}
EVIDENCE_FILL = {
    "OBSERVED": "E3EFE3", "PARTIAL": "FDF3DC",
    "INFERRED": "FBE4DC", "UNVERIFIABLE": "EFE7F2", "SDK": "F0EDE8",
}
PHASE_FILL = {"P1": "FDE8DF", "P2": "F0EDE8", "P3": "F7F5F2"}
REQ_FILL = {"required": "FBE4DC", "conditional": "FDF3DC", "optional": "F0EDE8"}


class Builder:
    def __init__(self, plan_path, out_path=None):
        self.path = Path(plan_path).resolve()
        self.base = self.path.parent
        self.data = json.loads(self.path.read_text())
        self.meta = self.data.get("meta", {})
        THEME.update(self.meta.get("theme", {}))
        self.out = Path(out_path) if out_path else self.base / (
            self.meta.get("output_name") or f"{self.slug(self.meta.get('product', 'tracking'))}_TrackingPlan.xlsx")
        self.wb = Workbook()
        self._subtitles = []
        self._tab_n = 1          # tab 1 is the read me, created separately
        self.screens = self._screens()
        self.gallery_row = {sid: 4 + i for i, sid in enumerate(sorted(self.screens))}
        self.thumbs = self._thumbs()
        self._fonts()

    # ---------------------------------------------------------------- setup
    @staticmethod
    def slug(s):
        return re.sub(r"[^A-Za-z0-9]+", "", str(s)) or "Plan"

    def _fonts(self):
        f = THEME["font"]
        self.F_TITLE = Font(name=f, size=16, bold=True, color=THEME["ink"])
        self.F_SUB = Font(name=f, size=10, color=THEME["grey"])
        self.F_H = Font(name=f, size=9, bold=True, color="FFFFFF")
        self.F_BODY = Font(name=f, size=9, color=THEME["ink"])
        self.F_GREY = Font(name=f, size=9, color=THEME["grey"])
        self.F_ACCENT = Font(name=f, size=9, bold=True, color=THEME["accent"])
        self.F_SECTION = Font(name=f, size=11, bold=True, color=THEME["accent"])
        self.F_LINK = Font(name=f, size=9, bold=True, color=THEME["accent"], underline="single")
        self.F_LABEL = Font(name=f, size=9, bold=True, color=THEME["ink"])
        thin = Side(style="thin", color=THEME["line"])
        self.BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.FILL_H = PatternFill("solid", fgColor=THEME["ink"])
        self.FILL_BAND = PatternFill("solid", fgColor=THEME["paper"])
        self.FILL_ACCENT = PatternFill("solid", fgColor=THEME["accent"])
        self.WRAP = Alignment(wrap_text=True, vertical="top")
        self.WRAP_C = Alignment(wrap_text=True, vertical="top", horizontal="center")

    def _screens(self):
        out, folder = {}, self.base / SCREEN_DIR
        if folder.is_dir():
            for f in sorted(folder.iterdir()):
                if f.suffix.lower() in (".png", ".jpg", ".jpeg") and f.name.upper().startswith("SCR-"):
                    sid = f.name.split("_")[0].upper()
                    label = f.name.split("_", 1)[-1].rsplit(".", 1)[0].replace("-", " ")
                    out[sid] = (f.name, label)
        return out

    def _thumbs(self):
        """Small JPEGs so the workbook carries its own evidence and needs no external files."""
        if not self.screens:
            return None
        try:
            from PIL import Image
        except ImportError:
            print("note: Pillow not installed, screenshots will not be embedded", file=sys.stderr)
            return None
        d = self.base / SCREEN_DIR / "_thumbs"
        d.mkdir(exist_ok=True)
        for fname, _ in self.screens.values():
            out = d / (fname.rsplit(".", 1)[0] + ".jpg")
            if not out.exists():
                im = Image.open(self.base / SCREEN_DIR / fname).convert("RGB")
                h = 460
                im.resize((int(im.width * h / im.height), h), Image.LANCZOS).save(
                    out, "JPEG", quality=72, optimize=True)
        return d

    # ---------------------------------------------------------------- helpers
    def sheet(self, title, subtitle="", number=True):  # noqa: D401
        """Tabs are numbered in creation order. Hardcoding the numbers leaves gaps like
        1, 2, 5 when a plan omits optional sections, which reads as a broken file."""
        if number:
            self._tab_n += 1
            title = f"{self._tab_n}. {title}"
        ws = self.wb.create_sheet(title[:31])
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws["A1"].font = self.F_TITLE
        if subtitle:
            ws["A2"] = subtitle
            ws["A2"].font = self.F_SUB
            ws["A2"].alignment = self.WRAP
            self._subtitles.append(ws)   # merged later, once columns are sized
        ws.append([])
        return ws

    def header(self, ws, headers, widths=None):
        r = ws.max_row + 1
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=i, value=h)
            c.font, c.fill, c.alignment, c.border = self.F_H, self.FILL_H, self.WRAP, self.BORDER
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        return r

    def row(self, ws, values, band=False, fonts=None):
        r = ws.max_row + 1
        for i, v in enumerate(values, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = (fonts or {}).get(i, self.F_BODY)
            c.alignment, c.border = self.WRAP, self.BORDER
            if band:
                c.fill = self.FILL_BAND
        return r

    def finish(self, ws, header_row, ncols):
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{ws.max_row}"

    def tint(self, ws, row, col, table, key, bold=False):
        if key in table:
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=table[key])
            c.font = Font(name=THEME["font"], size=9, bold=bold, color=THEME["ink"])

    @staticmethod
    def split_screens(text):
        if not text:
            return [], ""
        ids = [s.upper() for s in re.findall(r"SCR-\d+", str(text), re.I)]
        rest = re.sub(r"SCR-\d+", "", str(text), flags=re.I).strip(" ,;")
        return ids, rest

    def screen_links(self, ws, row, first_col, text, n=5):
        """Up to n individually clickable screen references.

        A cell holds one hyperlink, so each screen gets its own cell. Events with no
        screen show the stated reason instead of a blank, because a blank reads as an
        oversight while a reason reads as a decision.
        """
        ids, rest = self.split_screens(text)
        if not ids:
            c = ws.cell(row=row, column=first_col, value=rest or "")
            c.font, c.alignment = self.F_GREY, self.WRAP
            ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=first_col + n - 1)
            for k in range(n):
                ws.cell(row=row, column=first_col + k).border = self.BORDER
            return
        for k in range(n):
            c = ws.cell(row=row, column=first_col + k)
            c.border, c.alignment = self.BORDER, self.WRAP_C
            if k < len(ids):
                c.value = ids[k]
                if ids[k] in self.gallery_row:
                    c.hyperlink = Hyperlink(ref=c.coordinate,
                                            location=f"'{GALLERY_TAB}'!A{self.gallery_row[ids[k]]}")
                    c.font = self.F_LINK
                else:
                    c.font = self.F_BODY

    @staticmethod
    def joined(v):
        return "\n".join(v) if isinstance(v, list) else (v or "")

    # ---------------------------------------------------------------- tabs
    def tab_readme(self):
        ws = self.wb.active
        ws.title = "1. Read me"
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"{self.meta.get('product', 'Product')} analytics tracking plan"
        ws["A1"].font = self.F_TITLE
        ws["A2"] = self.meta.get("subtitle", "")
        ws["A2"].font = self.F_SUB
        ws.append([])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 120
        if self.meta.get("the_ask"):
            r = ws.max_row + 1
            c = ws.cell(row=r, column=1, value="START HERE")
            c.font = Font(name=THEME["font"], size=11, bold=True, color="FFFFFF")
            c.fill, c.alignment = self.FILL_ACCENT, self.WRAP
            b = ws.cell(row=r, column=2, value=self.meta["the_ask"])
            b.font = Font(name=THEME["font"], size=10, bold=True, color=THEME["ink"])
            b.fill = PatternFill("solid", fgColor=PHASE_FILL["P1"])
            b.alignment = self.WRAP
        for sec in self.data.get("governance", []):
            r = ws.max_row + 2
            ws.cell(row=r, column=1, value=sec.get("title", "")).font = self.F_SECTION
            for label, text in sec.get("rows", []):
                r = ws.max_row + 1
                a = ws.cell(row=r, column=1, value=label)
                a.font, a.alignment = self.F_LABEL, self.WRAP
                b = ws.cell(row=r, column=2, value=text)
                b.font, b.alignment = self.F_BODY, self.WRAP

    def tab_phase1(self):
        p1 = [e for e in self.data["events"] if e.get("phase") == "P1"]
        if not p1:
            return
        ws = self.sheet("Phase 1 build ask", self.meta.get("phase1_intro", ""))
        hr = self.header(ws, ["#", "Event", "Source", "Evidence", "What it is for", "Fires when",
                              "Screen 1", "Screen 2", "Screen 3", "Screen 4", "Screen 5",
                              "Where the work lands", "Parameters to send", "KPI / gap it serves"],
                         [5, 28, 9, 13, 42, 46, 10, 10, 10, 10, 10, 26, 58, 20])
        for n, ev in enumerate(p1, 1):
            params = "\n".join(
                (ev.get("packs") or []) + (ev.get("extra_params") or [])) or "none beyond the auto-collected context"
            rr = self.row(ws, [n, ev["name"], ev.get("source", ""), ev.get("evidence", ""),
                               ev.get("why_p1", ""), ev.get("trigger", ""),
                               None, None, None, None, None,
                               ev.get("work", ""), params, ev.get("kpi", "")],
                          band=(n % 2 == 0), fonts={2: self.F_ACCENT, 3: self.F_ACCENT, 14: self.F_GREY})
            self.screen_links(ws, rr, 7, ev.get("screens", ""))
            self.tint(ws, rr, 4, EVIDENCE_FILL, ev.get("evidence", ""), bold=True)
        self.finish(ws, hr, 14)
        if self.data.get("phase1_config"):
            r = ws.max_row + 2
            ws.cell(row=r, column=1, value="Also required, and not events").font = self.F_SECTION
        for label, text in self.data.get("phase1_config", []):
            r = ws.max_row + 1
            a = ws.cell(row=r, column=1, value=label)
            a.font, a.alignment = self.F_LABEL, self.WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            b = ws.cell(row=r, column=3, value=text)
            b.font, b.alignment = self.F_BODY, self.WRAP
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)

    def tab_roadmap(self):
        if not self.data.get("roadmap"):
            return
        ws = self.sheet("Roadmap",
                        "Phase 1 is the ask on the table. Phase 2 is earned by Phase 1 data raising a question. "
                        "Phase 3 is gated on something that has not happened yet.")
        hr = self.header(ws, ["Phase", "Scope", "Events", "Trigger to start", "What it buys you"],
                         [28, 60, 12, 38, 58])
        for i, row in enumerate(self.data["roadmap"]):
            rr = self.row(ws, row, band=(i % 2 == 1), fonts={1: self.F_ACCENT})
            m = re.search(r"[Pp]hase\s*([123])", str(row[0]))
            self.tint(ws, rr, 1, PHASE_FILL, f"P{m.group(1)}" if m else "")
        self.finish(ws, hr, 5)
        for key, title, headers in (("cuts", "What was cut, and why", ["Cut", "Reason"]),
                                    ("deletions", "Deleted events and what covers them now",
                                     ["Deleted event", "Covered instead by"])):
            if not self.data.get(key):
                continue
            r = ws.max_row + 2
            ws.cell(row=r, column=1, value=title).font = self.F_SECTION
            hr2 = self.header(ws, headers)
            ws.merge_cells(start_row=hr2, start_column=2, end_row=hr2, end_column=5)
            for i, row in enumerate(self.data[key]):
                rr = self.row(ws, list(row)[:2], band=(i % 2 == 1), fonts={1: self.F_GREY})
                ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=5)

    def tab_packs(self):
        if not self.data.get("packs"):
            return
        ws = self.sheet("Attribute packs",
                        "Reusable parameter bundles. Every event attaches one or more packs, and the pack defines the "
                        "exact parameters, types and values. Req: required means it must be present whenever the pack is "
                        "attached and its absence is a tracking bug; conditional means present only in the circumstance "
                        "the Notes name, and simply absent otherwise, never null-filled; optional means its absence "
                        "should never fail QA.")
        hr = self.header(ws, ["Pack", "Parameter", "Req", "Type", "Transport", "Values / example", "Notes"],
                         [22, 26, 13, 10, 20, 44, 60])
        for pack in self.data["packs"]:
            r = ws.max_row + 1
            c = ws.cell(row=r, column=1, value=f"{pack['id']}  ·  {pack.get('name', '')}")
            c.font = Font(name=THEME["font"], size=9, bold=True, color="FFFFFF")
            c.fill, c.alignment = self.FILL_ACCENT, self.WRAP
            for i in range(2, 8):
                cc = ws.cell(row=r, column=i, value=pack.get("purpose") if i == 2 else None)
                cc.fill = self.FILL_ACCENT
                cc.font = Font(name=THEME["font"], size=9, color="FFFFFF")
                cc.alignment = self.WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            for j, p in enumerate(pack.get("params", [])):
                rr = self.row(ws, [pack["id"], p["name"], p.get("req", ""), p.get("type", ""),
                                   p.get("transport", ""), p.get("values", ""), p.get("notes", "")],
                              band=(j % 2 == 1),
                              fonts={1: self.F_GREY, 2: self.F_ACCENT if p.get("register") else self.F_BODY})
                self.tint(ws, rr, 3, REQ_FILL, p.get("req", ""), bold=(p.get("req") == "required"))
        self.finish(ws, hr, 7)

    def tab_events(self):
        ws = self.sheet("All events",
                        "The complete target state, phased. Build Phase 1 first, on tab 2. Evidence: OBSERVED means seen "
                        "firing in the app; PARTIAL means the surface was seen but a branch was not exercised; INFERRED "
                        "means never seen; UNVERIFIABLE means it needs a real transaction or an unlaunched feature. "
                        "Treat anything not OBSERVED as a design to confirm, not a spec to implement. "
                        "Source: FE is client, BE is server, FE + BE means the client fires and the server backfills "
                        "only if it never arrives, SDK is automatic.")
        hr = self.header(ws, ["#", "Phase", "Source", "Evidence", "Event", "Journey", "Type", "Fires when",
                              "Screen 1", "Screen 2", "Screen 3", "Screen 4", "Screen 5",
                              "Platforms", "Collection", "Packs carried", "Event-specific parameters",
                              "Key event", "KPI link", "Notes"],
                         [5, 7, 9, 13, 28, 19, 12, 44, 10, 10, 10, 10, 10, 10, 24, 46, 36, 8, 14, 68])
        for n, ev in enumerate(self.data["events"], 1):
            packs = self.joined(ev.get("packs")) or (
                "auto-collected context only" if ev.get("type") == "auto" else "none beyond the auto-collected context")
            rr = self.row(ws, [n, ev.get("phase", ""), ev.get("source", ""), ev.get("evidence", ""),
                               ev["name"], ev.get("journey", ""), ev.get("type", ""), ev.get("trigger", ""),
                               None, None, None, None, None,
                               ev.get("platforms", ""), ev.get("collection", ""),
                               packs, self.joined(ev.get("extra_params")),
                               "KEY" if ev.get("key_event") else "", ev.get("kpi", ""), ev.get("notes", "")],
                          band=(n % 2 == 0),
                          fonts={3: self.F_ACCENT, 5: self.F_ACCENT if ev.get("key_event") else self.F_BODY,
                                 18: self.F_ACCENT})
            self.screen_links(ws, rr, 9, ev.get("screens", ""))
            self.tint(ws, rr, 2, PHASE_FILL, ev.get("phase", ""), bold=True)
            self.tint(ws, rr, 4, EVIDENCE_FILL, ev.get("evidence", ""))
        self.finish(ws, hr, 20)

    def tab_simple(self, title, subtitle, headers, widths, rows, accent_col=1):
        if not rows:
            return
        ws = self.sheet(title, subtitle)
        hr = self.header(ws, headers, widths)
        for i, row in enumerate(rows):
            self.row(ws, row, band=(i % 2 == 1), fonts={accent_col: self.F_ACCENT})
        self.finish(ws, hr, len(headers))

    def tab_user_props(self):
        ups = self.data.get("user_properties")
        if not ups:
            return
        ws = self.sheet("User properties",
                        "Persistent user-scoped attributes. These are what let every event be sliced by who the person is.")
        hr = self.header(ws, ["Phase", "Property", "Type", "Values", "Update rule", "Why it exists"],
                         [8, 24, 12, 34, 40, 60])
        for i, u in enumerate(ups):
            rr = self.row(ws, [u.get("phase", ""), u["name"], u.get("type", ""), u.get("values", ""),
                               u.get("update", ""), u.get("why", "")], band=(i % 2 == 1), fonts={2: self.F_ACCENT})
            self.tint(ws, rr, 1, PHASE_FILL, u.get("phase", ""), bold=True)
        self.finish(ws, hr, 6)

    def tab_definitions(self):
        """Registered dimensions and metrics, derived from the pack flags so they cannot drift."""
        packs = self.data.get("packs") or []
        if not packs:
            return
        vals = {p["name"]: p.get("values", "") for pk in packs for p in pk.get("params", [])}
        owner = {p["name"]: pk["id"] for pk in packs for p in pk.get("params", [])}
        dims = [p["name"] for pk in packs for p in pk.get("params", []) if p.get("register")]
        mets = [p["name"] for pk in packs for p in pk.get("params", []) if p.get("register_metric")]
        ups = self.data.get("user_properties") or []
        if not (dims or mets or ups):
            return
        lim = self.meta.get("limits", {})
        ws = self.sheet("Custom definitions",
                        "Event parameters stay invisible in most analytics interfaces until they are registered, and "
                        "registration is almost never backfilled. Register these before the instrumented release ships, "
                        "or every day before registration is blank for good.")
        for title, rows, headers, widths in (
            (f"Event-scoped dimensions ({len(dims)} of {lim.get('event_dimensions', 50)})",
             [[d, owner.get(d, ""), vals.get(d, "")] for d in dims], ["Parameter", "Pack", "Values"], [30, 14, 96]),
            (f"User-scoped dimensions ({len(ups)} of {lim.get('user_dimensions', 25)})",
             [[u["name"], u.get("phase", ""), u.get("values", "")] for u in ups],
             ["User property", "Phase", "Values"], [30, 14, 96]),
            (f"Custom metrics ({len(mets)} of {lim.get('metrics', 50)})",
             [[m, owner.get(m, ""), vals.get(m, "")] for m in mets], ["Parameter", "Pack", "Unit / values"], [30, 14, 96]),
        ):
            if not rows:
                continue
            r = ws.max_row + 2
            ws.cell(row=r, column=1, value=title).font = self.F_SECTION
            self.header(ws, headers, widths)
            for i, row in enumerate(rows):
                self.row(ws, row, band=(i % 2 == 1), fonts={1: self.F_ACCENT})
        ws.freeze_panes = "A4"

    def tab_gallery(self):
        if not self.screens:
            return
        # Not numbered: the internal links from the event tabs address this sheet by name,
        # so its title must stay exactly GALLERY_TAB.
        ws = self.sheet(GALLERY_TAB, number=False,
                        subtitle="Every screenshot from the walkthrough, embedded so the workbook needs no "
                                 "external files. Clicking a Screen link on the event tabs jumps here. This is "
                                 "also the reverse lookup: it says which events each screen evidences.")
        hr = self.header(ws, ["Screen", "What it shows", "Events it evidences", "File (full resolution)", "Screenshot"],
                         [12, 34, 50, 40, 22])
        uses = {}
        for ev in self.data["events"]:
            for sid in self.split_screens(ev.get("screens", ""))[0]:
                uses.setdefault(sid, []).append(ev["name"])
        for i, sid in enumerate(sorted(self.screens)):
            fname, label = self.screens[sid]
            rr = self.row(ws, [sid, label, ", ".join(uses.get(sid, [])) or "not referenced", fname, None],
                          band=(i % 2 == 1), fonts={1: self.F_ACCENT, 4: self.F_GREY})
            if self.thumbs:
                img = XLImage(str(self.thumbs / (fname.rsplit(".", 1)[0] + ".jpg")))
                img.width, img.height = 135, 300
                ws.add_image(img, f"E{rr}")
                ws.row_dimensions[rr].height = 232
        self.finish(ws, hr, 5)

    # ---------------------------------------------------------------- layout
    def autofit(self, ws, max_lines=28, line_pt=12.0):
        """Explicit row heights. Spreadsheet writers emit none, and each app then guesses
        differently, which is why long text looks clipped in one viewer and fine in another."""
        merged = {(r.min_row, r.min_col): r.max_col - r.min_col + 1 for r in ws.merged_cells.ranges}
        widths = {i: (ws.column_dimensions[get_column_letter(i)].width or 8.43)
                  for i in range(1, ws.max_column + 1)}
        for r in range(1, ws.max_row + 1):
            lines = 1
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=ci)
                v = cell.value
                if not isinstance(v, str) or not v:
                    continue
                if not (cell.alignment and cell.alignment.wrap_text):
                    continue  # titles overflow sideways; measuring them balloons the row
                span = merged.get((r, ci), 1)
                usable = sum(widths.get(ci + k, 0) for k in range(span)) - 1.5
                if usable <= 2:
                    continue
                lines = max(lines, sum(max(1, -(-len(seg) // int(usable))) for seg in v.split("\n")))
            if lines > 1:
                ws.row_dimensions[r].height = min(lines, max_lines) * line_pt + 4

    def wrap_subtitles(self):
        """A2 carries the paragraph that explains the tab. Left unmerged it runs off the
        page; merged across the used width it wraps and prints."""
        for ws in self._subtitles:
            last = max(2, min(ws.max_column, 8))
            try:
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last)
            except ValueError:
                pass

    def finalise(self):
        self.wrap_subtitles()
        single_table = {"2. Phase 1 build ask", "4. Attribute packs", "5. All events",
                        "6. User properties", GALLERY_TAB}
        for ws in self.wb.worksheets:
            if ws.title != GALLERY_TAB:
                self.autofit(ws)
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_margins.left = ws.page_margins.right = 0.3
            has_second_table = ws.title.endswith("Phase 1 build ask") and self.data.get("phase1_config")
            if ws.title in single_table and ws.freeze_panes and not has_second_table:
                m = re.match(r"[A-Z]+(\d+)", str(ws.freeze_panes))
                if m and int(m.group(1)) > 1:
                    ws.print_title_rows = f"{int(m.group(1)) - 1}:{int(m.group(1)) - 1}"

    def build(self):
        d = self.data
        self.tab_readme()
        self.tab_phase1()
        self.tab_roadmap()
        self.tab_packs()
        self.tab_events()
        self.tab_user_props()
        self.tab_definitions()
        self.tab_simple("Key events & audiences", "The conversions the platform optimises around, and the standing segments.",
                        ["Key event", "Why it is a key event", "KPI"], [30, 96, 20],
                        [[k["name"], k.get("why", ""), k.get("kpi", "")] for k in d.get("key_events", [])])
        self.tab_simple("Journey funnels", "The ordered event sequences behind every core funnel, with the measured baseline where one exists.",
                        ["Journey", "KPI", "Ordered events", "Baseline today"], [32, 12, 90, 56],
                        [[j["name"], j.get("kpi", ""), "  ->  ".join(j.get("events", [])), j.get("baseline", "")]
                         for j in d.get("journeys", [])])
        self.tab_simple("Gaps closed", "Every known problem, the evidence for it, and the event that closes it.",
                        ["Gap today", "Evidence", "Closed by"], [44, 56, 60],
                        [[g["gap"], g.get("evidence", ""), g.get("fix", "")] for g in d.get("gaps_closed", [])], 3)
        self.tab_simple("Journey map", "Screen-by-screen record of the walkthrough. Every trigger definition is grounded in a surface here.",
                        ["Screen / surface", "What the user sees and does", "Events fired", "Notes and findings"],
                        [26, 60, 44, 60], d.get("journey_map", []), 3)
        self.tab_simple("Confirm before build", "What is designed rather than observed. Settle these with engineering first; several are straight replacements of a placeholder enum with their real values.",
                        ["Item", "What the plan asserts, and why it is uncertain", "What to ask for"], [30, 74, 66],
                        d.get("open_questions", []))
        self.tab_gallery()
        self.finalise()
        self.wb.save(self.out)
        return self.out


def main():
    ap = argparse.ArgumentParser(description="Build a tracking-plan workbook from a plan JSON file.")
    ap.add_argument("plan", help="path to plan.json")
    ap.add_argument("-o", "--output", help="output .xlsx path")
    a = ap.parse_args()
    b = Builder(a.plan, a.output)
    out = b.build()
    ev = b.data["events"]
    print(f"Saved {out} ({out.stat().st_size:,} bytes)")
    print(f"  {len(ev)} events, {len(b.wb.sheetnames)} tabs, {len(b.screens)} screens embedded")


if __name__ == "__main__":
    main()
